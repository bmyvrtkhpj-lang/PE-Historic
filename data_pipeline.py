"""
data_pipeline.py -- as shared by user, with dtype-coercion fix and Ticker.history()
"""
import pandas as pd
from openpyxl import load_workbook


def find_section_row(ws, label, max_row=250):
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None and str(val).strip().lower() == label.strip().lower():
            return r
    return None


def _row_values(ws, row):
    vals = []
    c = 2
    while True:
        v = ws.cell(row=row, column=c).value
        if v is None:
            break
        vals.append(v)
        c += 1
    return vals


def extract_annual_fundamentals(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Data Sheet"]
    company_name = ws["B1"].value

    pl_header_row = find_section_row(ws, "PROFIT & LOSS")
    quarters_header_row = find_section_row(ws, "Quarters")
    if pl_header_row is None or quarters_header_row is None:
        raise ValueError("Could not locate 'PROFIT & LOSS' / 'Quarters' section headers.")

    report_date_row = None
    net_profit_row = None
    for r in range(pl_header_row, quarters_header_row):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        label = str(label).strip()
        if label == "Report Date" and report_date_row is None:
            report_date_row = r
        if label == "Net profit" and net_profit_row is None:
            net_profit_row = r

    if report_date_row is None or net_profit_row is None:
        raise ValueError("Could not find annual 'Report Date' / 'Net profit' rows.")

    derived_row = find_section_row(ws, "DERIVED:")
    if derived_row is None:
        raise ValueError("Could not find 'DERIVED:' section.")

    shares_row = None
    for r in range(derived_row, derived_row + 10):
        label = ws.cell(row=r, column=1).value
        if label and "Adjusted Equity Shares" in str(label):
            shares_row = r
            break
    if shares_row is None:
        raise ValueError("Could not find 'Adjusted Equity Shares in Cr' row.")

    dates = _row_values(ws, report_date_row)
    net_profit = _row_values(ws, net_profit_row)
    shares = _row_values(ws, shares_row)

    n = min(len(dates), len(net_profit), len(shares))
    df = pd.DataFrame({
        "fiscal_year_end": pd.to_datetime(dates[:n]),
        "net_profit_cr": net_profit[:n],
        "shares_cr": shares[:n],
    })
    df["annual_eps"] = df["net_profit_cr"] / df["shares_cr"]
    df["yoy_shares_change_pct"] = df["shares_cr"].pct_change() * 100
    df["company"] = company_name
    return df


def build_step_function_pe(annual_df, daily_price, filing_lag_days=60):
    eps_df = annual_df.copy()
    eps_df["effective_date"] = pd.to_datetime(eps_df["fiscal_year_end"]) + pd.Timedelta(days=filing_lag_days)
    eps_df["effective_date"] = pd.to_datetime(eps_df["effective_date"]).dt.tz_localize(None).astype("datetime64[ns]")
    eps_df = eps_df.sort_values("effective_date")

    price_df = daily_price.rename("price").to_frame().reset_index()
    price_df.columns = ["date", "price"]
    price_df["date"] = pd.to_datetime(price_df["date"]).dt.tz_localize(None).astype("datetime64[ns]")

    merged = pd.merge_asof(
        price_df.sort_values("date"),
        eps_df[["effective_date", "annual_eps"]].rename(columns={"effective_date": "date"}),
        on="date",
        direction="backward",
    )
    merged["pe"] = merged["price"] / merged["annual_eps"]
    return merged.set_index("date")

def apply_corporate_action_exclusions(pe_df, exclusions):
    out = pe_df.copy()
    for start, end in exclusions or []:
        out.loc[start:end, "pe"] = None
    return out


def fetch_price_volume(ticker, start, end):
    import yfinance as yf
    stock = yf.Ticker(ticker)
    raw = stock.history(start=start, end=end)
    if raw.empty:
        return pd.DataFrame(columns=["price", "volume"])
    out = pd.DataFrame({
        "price": pd.to_numeric(raw["Close"], errors="coerce"),
        "volume": pd.to_numeric(raw["Volume"], errors="coerce"),
    }).dropna()
    out.index = pd.to_datetime(out.index).tz_localize(None)
    return out


def load_eod2_delivery_data(csv_path):
    """
    Loads a LOCAL per-stock CSV (if you cloned eod2_data yourself).
    For the deployed app, use fetch_eod2_delivery_data() instead, which
    fetches directly over HTTP without needing a git clone at all.
    """
    df = pd.read_csv(csv_path, parse_dates=["Date"]).set_index("Date")
    return _process_eod2_dataframe(df)


def fetch_eod2_delivery_data(nse_symbol):
    """
    Fetches a stock's price/volume/delivery CSV DIRECTLY from EOD2's public
    data repo over HTTP -- no git clone needed. Confirmed working (tested
    live): raw.githubusercontent.com serves individual files from the repo
    without cloning the whole ~3500-file repo, which matters for Streamlit
    Cloud (a full clone would be slow/wasteful for a handful of stocks).

    nse_symbol: the official NSE trading symbol (e.g. "HDFCBANK", "M&M",
    "BAJAJ-AUTO") -- NOT the yfinance-style ".NS" ticker. Confirmed the
    repo's filenames are simply nse_symbol.lower() + ".csv", special
    characters (&, -) included as-is, no URL-encoding needed (tested).

    Returns the same processed DataFrame as load_eod2_delivery_data(), or
    an empty DataFrame with a clear reason if the symbol isn't found.
    """
    import requests
    import io as _io

    filename = f"{nse_symbol.lower()}.csv"
    url = f"https://raw.githubusercontent.com/BennyThadikaran/eod2_data/main/daily/{filename}"
    try:
        res = requests.get(url, timeout=15)
    except Exception as exc:
        return pd.DataFrame(), f"Network error fetching {nse_symbol}: {exc}"

    if res.status_code != 200:
        return pd.DataFrame(), f"{nse_symbol}: not found in eod2_data (HTTP {res.status_code}). Check the NSE symbol spelling."

    df = pd.read_csv(_io.StringIO(res.text), parse_dates=["Date"]).set_index("Date")
    return _process_eod2_dataframe(df), None


def _process_eod2_dataframe(df):
    """
    Shared processing for EOD2 data (local or fetched).
    Now keeps Open, High, Low, Close for Candlestick charting.
    """
    df["delivery_pct"] = df["DLV_QTY"] / df["Volume"] * 100
    bad = df["delivery_pct"] > 100
    n_bad = int(bad.sum())
    if n_bad:
        df.loc[bad, "delivery_pct"] = None

    # ADDED: Open, High, Low columns for Candlestick charts
    out = df[["Open", "High", "Low", "Close", "Volume", "DLV_QTY", "delivery_pct"]].rename(
        columns={"Close": "price", "Volume": "volume", "Open": "open", "High": "high", "Low": "low"}
    )
    out = out.dropna(subset=["price"])
    out.attrs["n_impossible_delivery_pct_rows_removed"] = n_bad
    return out
